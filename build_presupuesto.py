# -*- coding: utf-8 -*-
"""Genera el presupuesto profesional del muro perimetral (mampostería confinada)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# ----------------------------------------------------------------------------
# Paleta (skill excel-formato-pro)
AZUL      = "1F4E79"
GRIS_SUB  = "D9D9D9"
GRIS_TOT  = "595959"
BORDE     = "BFBFBF"
PAST_MURO = "DDEBF7"
PAST_PORT = "E2EFDA"
PAST_MO   = "FCE4D6"
AZUL_EDIT = "0000FF"

thin = Side(style="thin", color=BORDE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FUENTE = "Calibri"
def F(sz=11, bold=False, color="000000", italic=False):
    return Font(name=FUENTE, size=sz, bold=bold, color=color, italic=italic)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
LEFTT  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

MONEDA = '"C$"#,##0.00'
CANT   = '#,##0.##'

wb = Workbook()

# ============================================================================
# HOJA 1: PRESUPUESTO
# ============================================================================
ws = wb.active
ws.title = "Presupuesto"
ws.sheet_view.showGridLines = False

anchos = {"A": 7, "B": 50, "C": 11, "D": 11, "E": 16, "F": 17}
for col, w in anchos.items():
    ws.column_dimensions[col].width = w

def merge_set(rng, value, font=None, fill=None, align=None, fmt=None, border=False):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = value
    if font:  c.font = font
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    mc, mr, xc, xr = range_boundaries(rng)
    for r in range(mr, xr + 1):
        for cc in range(mc, xc + 1):
            cel = ws.cell(row=r, column=cc)
            if fill:   cel.fill = PatternFill("solid", fgColor=fill)
            if border: cel.border = BORDER
    return c

# ---- Encabezado tipo cotización -------------------------------------------
merge_set("A1:F1", "PRESUPUESTO DE MATERIALES Y MANO DE OBRA",
          F(16, True, AZUL), align=CENTER)
ws.row_dimensions[1].height = 26
merge_set("A2:F2",
          "Elevación de muro perimetral — Mampostería confinada de bloque de 6\" sobre muro de cantera existente",
          F(11, False, "404040", italic=True), align=CENTER)
ws.row_dimensions[2].height = 20

hoy = datetime.date(2026, 8, 4)
info = [
    ("Documento N°:",         "PRE-2026-001"),
    ("Fecha de emisión:",     hoy),
    ("Válida hasta:",         "=C5+30"),
    ("Proyecto / Ubicación:", "Muro perimetral — Managua, Nicaragua"),
    ("Moneda / Impuesto:",    "Córdoba (C$) — Precios con IVA 15% incluido"),
]
r = 4
for etiqueta, valor in info:
    merge_set(f"A{r}:B{r}", etiqueta, F(11, True), align=RIGHT)
    cval = merge_set(f"C{r}:F{r}", valor, F(11), align=LEFT)
    if etiqueta.startswith("Fecha") or etiqueta.startswith("Válida"):
        cval.number_format = "dd/mm/yyyy"
        cval.alignment = LEFT
    ws.row_dimensions[r].height = 19
    r += 1

# ---- Utilidades de tabla ---------------------------------------------------
HEAD = ["Ítem", "Descripción", "Unidad", "Cantidad", "Precio Unitario (C$)", "Total (C$)"]

def escribir_encabezado(row):
    for i, h in enumerate(HEAD):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = F(11, True, "FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 36

def escribir_seccion(row, texto, pastel):
    merge_set(f"A{row}:F{row}", texto, F(12, True, AZUL), fill=pastel,
              align=Alignment(horizontal="left", vertical="center"), border=True)
    ws.row_dimensions[row].height = 24

def escribir_items(start, items, item_offset=0):
    row = start
    for idx, (desc, uni, cant, pu) in enumerate(items, start=1):
        n = item_offset + idx
        ws.cell(row=row, column=1, value=n).font = F(11)
        ws.cell(row=row, column=2, value=desc).font = F(11)
        ws.cell(row=row, column=3, value=uni).font = F(11)
        cc = ws.cell(row=row, column=4, value=cant); cc.font = F(11, color=AZUL_EDIT); cc.number_format = CANT
        ce = ws.cell(row=row, column=5, value=pu);   ce.font = F(11, color=AZUL_EDIT); ce.number_format = MONEDA
        cf = ws.cell(row=row, column=6, value=f"=D{row}*E{row}"); cf.font = F(11); cf.number_format = MONEDA
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=5).alignment = RIGHT
        ws.cell(row=row, column=6).alignment = RIGHT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        ws.row_dimensions[row].height = 21
        row += 1
    return row

def fila_subtotal(row, etiqueta, formula):
    merge_set(f"A{row}:E{row}", etiqueta, F(11, True), fill=GRIS_SUB, align=RIGHT, border=True)
    c = ws.cell(row=row, column=6, value=formula)
    c.font = F(11, True); c.number_format = MONEDA; c.fill = PatternFill("solid", fgColor=GRIS_SUB)
    c.alignment = RIGHT; c.border = BORDER
    ws.row_dimensions[row].height = 22

# ---- Datos -----------------------------------------------------------------
materiales_muro = [
    ("Bloque de concreto de 6\" (15x20x40 cm)",          "Unidad", 80,  24.00),
    ("Varilla corrugada de 3/8\" G-40 (barra de 6 m)",   "Unidad", 79,  195.00),
    ("Estribo de 1/4\" armado de 10x10 cm",              "Unidad", 100, 15.00),
    ("Alambre de amarre recocido cal. 18",               "Libra",  4,   28.00),
    ("Cemento gris tipo GU (bolsa de 42.5 kg)",          "Bolsa",  5,   385.00),
    ("Arena de construcción",                            "m³",     1,   800.00),
    ("Piedrín de 1/2\"",                                 "m³",     0.5, 1000.00),
    ("Tabla de pino de 1x12x6 varas",                    "Unidad", 3,   520.00),
    ("Regla de pino de 1x2x6 varas",                     "Unidad", 2,   110.00),
    ("Clavos corrientes de 2 1/2\"",                     "Libra",  5,   35.00),
    ("Clavos de acero de 2 1/2\"",                       "Docena", 2,   55.00),
    ("Adhesivo Plaster Bond (concreto-concreto)",        "Litro",  1,   230.00),
]
materiales_porton = [
    ("Rollo de concertina / serpentina de 450 mm",       "Rollo",  3,  2300.00),
    ("Grapas para cerca",                                "Libra",  4,  35.00),
    ("Electrodo de soldadura 6011 de 3/32\" (caja)",     "Caja",   1,  780.00),
    ("Pintura anticorrosiva",                            "Cuarto", 2,  180.00),
    ("Diluyente (thinner corriente)",                    "Galón",  1,  400.00),
    ("Tubo cuadrado de 1 1/2\" chapa 14 (6 m)",          "Unidad", 3,  580.00),
    ("Lámina negra de 4x8 pies x 1/16\"",                "Unidad", 1,  1600.00),
    ("Par de bisagra de bala",                           "Par",    1,  120.00),
    ("Pasador (aldaba) de 5\"",                          "Unidad", 1,  90.00),
    ("Disco de corte fino de 7\" para metal",            "Unidad", 2,  75.00),
]
mano_obra = [
    ("Mano de obra — elevación de muro perimetral",      "Global", 1,  8000.00),
    ("Mano de obra — concertina y portón metálico",      "Global", 1,  7500.00),
]

escribir_seccion(10, "1.  MATERIALES PARA MURO (MAMPOSTERÍA CONFINADA)", PAST_MURO)
escribir_encabezado(11)
fin1 = escribir_items(12, materiales_muro, item_offset=0)
sub1 = fin1
fila_subtotal(sub1, "Subtotal materiales — muro", f"=SUM(F12:F{fin1-1})")

sec2 = sub1 + 1
escribir_seccion(sec2, "2.  MATERIALES PARA PORTÓN, CERCO Y CONCERTINA", PAST_PORT)
escribir_encabezado(sec2 + 1)
ini2 = sec2 + 2
fin2 = escribir_items(ini2, materiales_porton, item_offset=len(materiales_muro))
sub2 = fin2
fila_subtotal(sub2, "Subtotal materiales — portón y concertina", f"=SUM(F{ini2}:F{fin2-1})")

sec3 = sub2 + 1
escribir_seccion(sec3, "3.  MANO DE OBRA", PAST_MO)
escribir_encabezado(sec3 + 1)
ini3 = sec3 + 2
fin3 = escribir_items(ini3, mano_obra, item_offset=len(materiales_muro) + len(materiales_porton))
sub3 = fin3
fila_subtotal(sub3, "Subtotal mano de obra", f"=SUM(F{ini3}:F{fin3-1})")

# ---- Resumen ---------------------------------------------------------------
res = sub3 + 2
escribir_seccion(res, "RESUMEN DEL PRESUPUESTO", "F2F2F2")
def fila_resumen(row, etiqueta, formula, fmt=MONEDA, fill=None, blanco=False, bold=True, big=False):
    merge_set(f"A{row}:E{row}", etiqueta, F(13 if big else 11, bold, "FFFFFF" if blanco else "000000"),
              fill=fill, align=RIGHT, border=True)
    c = ws.cell(row=row, column=6, value=formula)
    c.font = F(13 if big else 11, bold, "FFFFFF" if blanco else "000000")
    c.number_format = fmt; c.alignment = RIGHT; c.border = BORDER
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[row].height = 26 if big else 22

r_mat  = res + 1
r_mo   = res + 2
r_subg = res + 3
r_imp  = res + 4
r_tot  = res + 5
fila_resumen(r_mat,  "Subtotal materiales (IVA incluido)", f"=F{sub1}+F{sub2}", fill=GRIS_SUB, bold=False)
fila_resumen(r_mo,   "Subtotal mano de obra",              f"=F{sub3}",         fill=GRIS_SUB, bold=False)
fila_resumen(r_subg, "Subtotal general",                   f"=F{r_mat}+F{r_mo}", fill=GRIS_SUB)
fila_resumen(r_imp,  "Imprevistos (5%)",                   f"=F{r_subg}*0.05",   fill=GRIS_SUB)
fila_resumen(r_tot,  "TOTAL GENERAL (C$)",                 f"=F{r_subg}+F{r_imp}", fill=GRIS_TOT, blanco=True, big=True)

r_iva = r_tot + 1
merge_set(f"A{r_iva}:E{r_iva}", "IVA (15%) ya contenido en el monto de materiales:",
          F(9, False, "595959", italic=True), align=RIGHT)
merge_set(f"F{r_iva}:F{r_iva}", f"=(F{sub1}+F{sub2})/1.15*0.15",
          F(9, False, "595959", italic=True), align=RIGHT, fmt=MONEDA)
ws.row_dimensions[r_iva].height = 16
r_usd = r_iva + 1
merge_set(f"A{r_usd}:E{r_usd}", "Total general equivalente en dólares (referencial, no vinculante):",
          F(9, False, "595959", italic=True), align=RIGHT)
merge_set(f"F{r_usd}:F{r_usd}", f"=F{r_tot}/36.62",
          F(9, False, "595959", italic=True), align=RIGHT, fmt='"US$ "#,##0.00')
ws.row_dimensions[r_usd].height = 16

# ---- Observaciones ---------------------------------------------------------
obs_title = r_usd + 2
escribir_seccion(obs_title, "OBSERVACIONES", "F2F2F2")
observaciones = [
    "1. Precios unitarios según el mercado de ferreterías de Managua, Nicaragua (agosto 2026); incluyen IVA (15%). Están sujetos a variación por disponibilidad y fluctuación del mercado.",
    "2. El muro se levanta sobre un muro de cantera existente y se continúa con bloque de 6\" en mampostería confinada (columnas y viga corona de amarre con refuerzo de 3/8\" y estribos de 1/4\").",
    "3. Las cantidades provienen de la lista entregada por el cliente. No se incluye desperdicio adicional al 5% de imprevistos ni acarreo de materiales fuera del sitio.",
    "4. No incluye: excavación o refuerzo de cimiento nuevo, trazo topográfico, permisos municipales, ni acabado fino (repello/afinado), salvo indicación contraria.",
    "5. La mano de obra es por trato (precio global) con maestro de obra; no incluye prestaciones de ley ni alimentación, salvo acuerdo por escrito.",
    "6. La madera (tablas y reglas) es para formaleta/encofrado de columnas y viga; es material reutilizable no consumido en su totalidad.",
    "7. Validez de la oferta: 30 días calendario a partir de la fecha de emisión. Tipo de cambio referencial: 1 USD ≈ C$ 36.62.",
]
ro = obs_title + 1
for o in observaciones:
    merge_set(f"A{ro}:F{ro}", o, F(10), align=LEFTT, border=True)
    lineas = max(1, int(len(o) / 95) + 1)
    ws.row_dimensions[ro].height = 15 * lineas + 6
    ro += 1

# ---- Leyenda ---------------------------------------------------------------
ley = ro + 1
merge_set(f"A{ley}:F{ley}", "LEYENDA DE COLORES", F(11, True, AZUL), align=LEFT)
ws.row_dimensions[ley].height = 20
leyendas = [
    ("Texto azul", AZUL_EDIT, "Celdas editables por el usuario (cantidad y precio unitario). El total se calcula automáticamente."),
    ("Fondo gris claro", GRIS_SUB, "Subtotales y resumen del presupuesto."),
    ("Fondo gris oscuro", GRIS_TOT, "Total general del presupuesto."),
]
rl = ley + 1
for etq, color, desc in leyendas:
    c = ws.cell(row=rl, column=1, value=etq)
    c.fill = PatternFill("solid", fgColor="FFFFFF" if color == AZUL_EDIT else color)
    c.font = F(10, True, AZUL_EDIT if color == AZUL_EDIT else ("FFFFFF" if color == GRIS_TOT else "000000"))
    c.alignment = CENTER; c.border = BORDER
    ws.merge_cells(f"B{rl}:F{rl}")
    d = ws[f"B{rl}"]; d.value = desc; d.font = F(10); d.alignment = LEFT; d.border = BORDER
    for cc in range(1, 7):
        ws.cell(row=rl, column=cc).border = BORDER
    ws.row_dimensions[rl].height = 18
    rl += 1

# ---- Firmas ----------------------------------------------------------------
fir = rl + 2
top = Side(style="thin", color="000000")
ws.merge_cells(f"B{fir}:C{fir}")
ws.merge_cells(f"E{fir}:F{fir}")
for rng, txt in ((f"B{fir}:C{fir}", "Elaborado por"), (f"E{fir}:F{fir}", "Aprobado por / Cliente")):
    a = ws[rng.split(":")[0]]
    a.value = txt; a.font = F(11, True); a.alignment = CENTER
    mc, mr, xc, xr = range_boundaries(rng)
    for cc in range(mc, xc + 1):
        ws.cell(row=mr, column=cc).border = Border(top=top)
ws.row_dimensions[fir].height = 20
ws.row_dimensions[fir - 1].height = 34

# Validación de datos: lista de unidades
unidades = '"Unidad,Libra,Bolsa,Docena,Caja,Rollo,Cuarto,Galón,Par,m³,Global,Litro"'
dv = DataValidation(type="list", formula1=unidades, allow_blank=True)
ws.add_data_validation(dv)
for (ini, fin) in [(12, fin1 - 1), (ini2, fin2 - 1), (ini3, fin3 - 1)]:
    dv.add(f"C{ini}:C{fin}")

# Impresión / vista
ws.freeze_panes = "A12"
ws.print_area = f"A1:F{fir}"
ws.print_title_rows = "11:11"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5
ws.print_options.horizontalCentered = True
ws.sheet_view.zoomScale = 100

# ============================================================================
# HOJA 2: ESPECIFICACIONES TÉCNICAS
# ============================================================================
es = wb.create_sheet("Especificaciones Técnicas")
es.sheet_view.showGridLines = False
for col, w in {"A": 6, "B": 26, "C": 58, "D": 22}.items():
    es.column_dimensions[col].width = w

def es_merge(rng, value, font=None, fill=None, align=None):
    es.merge_cells(rng)
    c = es[rng.split(":")[0]]
    c.value = value
    if font: c.font = font
    if align: c.alignment = align
    mc, mr, xc, xr = range_boundaries(rng)
    for r in range(mr, xr + 1):
        for cc in range(mc, xc + 1):
            if fill: es.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=fill)
    return c

es_merge("A1:D1", "CUADRO DE ESPECIFICACIONES TÉCNICAS", F(16, True, AZUL), align=CENTER)
es.row_dimensions[1].height = 26
es_merge("A2:D2", "Muro perimetral de mampostería confinada sobre muro de cantera existente",
         F(11, False, "404040", italic=True), align=CENTER)
es.row_dimensions[2].height = 20

for i, h in enumerate(["N°", "Elemento / Partida", "Especificación técnica", "Norma / Referencia"]):
    c = es.cell(row=4, column=1 + i, value=h)
    c.font = F(11, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = CENTER; c.border = BORDER
es.row_dimensions[4].height = 32

specs = [
    ("Muro existente (base)", "Muro de cantera existente que sirve de base y arranque. Verificar plomo, nivel y estado antes de continuar; picar y limpiar la corona y aplicar puente de adherencia (Plaster Bond) antes de asentar la primera hilada de bloque.", "RNC-07 / buena práctica"),
    ("Unidad de mampostería", "Bloque de concreto hueco de 6\" (nominal 15x20x40 cm), asentado a soga con junta de 1.0 a 1.5 cm. Resistencia mínima recomendada f'm según uso estructural.", "ASTM C90 / NTON 12 008-09"),
    ("Mortero de pega", "Mortero cemento-arena en proporción 1:4 (tipo S). Espesor de junta 10-15 mm, llenas y enrasadas. Agua potable; arena cribada, libre de materia orgánica.", "ASTM C270"),
    ("Concreto de relleno / columnas", "Concreto en celdas reforzadas y columnas de confinamiento, proporción 1:2:2 (cemento:arena:piedrín de 1/2\"), resistencia f'c ≥ 3000 psi (210 kg/cm²). Vibrado o chuzeado.", "ACI 318 / RNC-07"),
    ("Refuerzo vertical", "Varilla corrugada de 3/8\" grado 40 en columnas de confinamiento y celdas, según separación de diseño. Traslapes mínimos de 40 diámetros; anclaje al muro de cantera y a la viga corona.", "ASTM A615 Gr.40"),
    ("Refuerzo transversal", "Estribos de 1/4\" de 10x10 cm, separación típica cada 15-20 cm en columnas y confinamiento de viga. Ganchos a 135°.", "ACI 318 / RNC-07"),
    ("Viga corona / solera", "Viga de amarre (corona) de concreto reforzado en la parte superior del muro, ligada a las columnas de confinamiento para dar rigidez al conjunto.", "RNC-07 mampostería confinada"),
    ("Amarres y fijación", "Alambre de amarre recocido cal. 18 para el armado del acero. Clavos corrientes y de acero de 2 1/2\" para la formaleta de madera (tabla y regla de pino).", "Práctica constructiva"),
    ("Encofrado / formaleta", "Formaleta de madera de pino (tabla 1x12 y regla 1x2) para columnas y viga; apuntalada, aplomada y humedecida antes del colado. Material reutilizable.", "Práctica constructiva"),
    ("Estructura del portón", "Marco de tubo cuadrado de 1 1/2\" chapa 14 con forro de lámina negra de 1/16\" (4x8 pies). Uniones soldadas con electrodo 6011 de 3/32\".", "AWS D1.1 / práctica de taller"),
    ("Herrajes del portón", "Par de bisagra de bala para el giro y pasador (aldaba) de 5\" para el cierre. Ajuste y lubricación final.", "Ferretería estándar"),
    ("Protección anticorrosiva", "Limpieza de escoria y desengrase; una mano de pintura anticorrosiva (base) diluida con thinner sobre el portón y los elementos metálicos.", "SSPC-SP2 / ficha técnica"),
    ("Concertina / serpentina", "Concertina de 450 mm fijada sobre la corona del muro con grapas y soportes, para seguridad perimetral. Tensado uniforme y anclaje firme.", "Práctica de seguridad"),
    ("Corte de metal", "Corte de tubo y lámina con disco de corte fino de 7\" para acero, con equipo y equipo de protección personal adecuado.", "Seguridad ocupacional"),
    ("Curado", "Curado del concreto y del mortero por humedecimiento durante un mínimo de 7 días, protegido del sol y el viento directo, para alcanzar la resistencia de diseño.", "ACI 308 / RNC-07"),
]
rs = 5
for i, (elem, espec, norma) in enumerate(specs, start=1):
    es.cell(row=rs, column=1, value=i).font = F(10)
    es.cell(row=rs, column=2, value=elem).font = F(10, True)
    es.cell(row=rs, column=3, value=espec).font = F(10)
    es.cell(row=rs, column=4, value=norma).font = F(10)
    es.cell(row=rs, column=1).alignment = CENTER
    es.cell(row=rs, column=2).alignment = LEFTT
    es.cell(row=rs, column=3).alignment = LEFTT
    es.cell(row=rs, column=4).alignment = LEFTT
    for cc in range(1, 5):
        es.cell(row=rs, column=cc).border = BORDER
    lineas = max(1, int(len(espec) / 72) + 1)
    es.row_dimensions[rs].height = 15 * lineas + 8
    if i % 2 == 0:
        for cc in range(1, 5):
            es.cell(row=rs, column=cc).fill = PatternFill("solid", fgColor="F2F2F2")
    rs += 1

nota = rs + 1
es_merge(f"A{nota}:D{nota}",
         "Nota: Este cuadro es de carácter referencial y de buena práctica constructiva. Las secciones, separaciones de refuerzo y resistencias definitivas deben ser validadas por el ingeniero responsable según el Reglamento Nacional de la Construcción (RNC-07) de Nicaragua.",
         F(9, italic=True, color="595959"), align=LEFTT)
es.row_dimensions[nota].height = 46

es.freeze_panes = "A5"
es.print_area = f"A1:D{nota}"
es.print_title_rows = "4:4"
es.page_setup.orientation = "landscape"
es.page_setup.fitToWidth = 1
es.page_setup.fitToHeight = 0
es.sheet_properties.pageSetUpPr.fitToPage = True
es.page_margins.left = es.page_margins.right = 0.4
es.page_margins.top = es.page_margins.bottom = 0.5
es.print_options.horizontalCentered = True

wb.active = 0
try:
    wb.calculation.fullCalcOnLoad = True
except Exception as e:
    print("aviso fullCalcOnLoad:", e)

ARCH = "Presupuesto_Muro_Perimetral.xlsx"
wb.save(ARCH)

# ---------------------------------------------------------------------------
# Inyección de valores en caché (LibreOffice no disponible en el entorno).
# ---------------------------------------------------------------------------
val_map = {}
r = 12
for (_d, _u, cant, pu) in materiales_muro:
    val_map[f"F{r}"] = round(cant * pu, 2); r += 1
val_map[f"F{sub1}"] = round(sum(c * p for _a, _b, c, p in materiales_muro), 2)
r = ini2
for (_d, _u, cant, pu) in materiales_porton:
    val_map[f"F{r}"] = round(cant * pu, 2); r += 1
val_map[f"F{sub2}"] = round(sum(c * p for _a, _b, c, p in materiales_porton), 2)
r = ini3
for (_d, _u, cant, pu) in mano_obra:
    val_map[f"F{r}"] = round(cant * pu, 2); r += 1
val_map[f"F{sub3}"] = round(sum(c * p for _a, _b, c, p in mano_obra), 2)

sub1v = val_map[f"F{sub1}"]; sub2v = val_map[f"F{sub2}"]; sub3v = val_map[f"F{sub3}"]
matv = round(sub1v + sub2v, 2); mov = sub3v
subgv = round(matv + mov, 2); impv = round(subgv * 0.05, 2); totv = round(subgv + impv, 2)
ivav = round((sub1v + sub2v) / 1.15 * 0.15, 2); usdv = round(totv / 36.62, 2)
val_map[f"F{r_mat}"] = matv
val_map[f"F{r_mo}"] = mov
val_map[f"F{r_subg}"] = subgv
val_map[f"F{r_imp}"] = impv
val_map[f"F{r_tot}"] = totv
val_map[f"F{r_iva}"] = ivav
val_map[f"F{r_usd}"] = usdv
val_map["C6"] = (datetime.date(2026, 9, 3) - datetime.date(1899, 12, 30)).days

import zipfile, re, os
with zipfile.ZipFile(ARCH) as zin:
    names = zin.namelist()
    blobs = {n: zin.read(n) for n in names}
xml = blobs["xl/worksheets/sheet1.xml"].decode("utf-8")
for coord, val in val_map.items():
    sval = str(int(val)) if float(val) == int(val) else f"{val:.2f}"
    pat = re.compile(rf'(<c r="{coord}"[^>]*><f>.*?</f>)<v\s*/>(</c>)', re.DOTALL)
    xml, n = pat.subn(rf'\g<1><v>{sval}</v>\g<2>', xml)
    if n == 0:
        print("WARN sin match:", coord)
blobs["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
tmp = ARCH + ".tmp"
with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, blobs[n])
os.replace(tmp, ARCH)

print("OK. sub1:", sub1v, "sub2:", sub2v, "sub3:", sub3v,
      "| subtotal general:", subgv, "imprevistos:", impv, "TOTAL:", totv,
      "| USD:", usdv)
