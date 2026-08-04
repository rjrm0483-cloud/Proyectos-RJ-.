# -*- coding: utf-8 -*-
"""Comparativo de precios: presupuesto del 30-jul-2026 vs. versión actual."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
import datetime

AZUL="1F4E79"; GRIS_SUB="D9D9D9"; GRIS_TOT="595959"; BORDE="BFBFBF"
VERDE="008000"; ROJO="C00000"; GRIST="595959"
PAST_JUL="FCE4D6"; PAST_ACT="E2EFDA"
thin=Side(style="thin",color=BORDE); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def F(sz=11,bold=False,color="000000",italic=False):
    return Font(name="Calibri",size=sz,bold=bold,color=color,italic=italic)
CENTER=Alignment("center",vertical="center",wrap_text=True)
LEFT=Alignment("left",vertical="center",wrap_text=True)
RIGHT=Alignment("right",vertical="center",wrap_text=True)
LEFTT=Alignment("left",vertical="top",wrap_text=True)
MON='"C$"#,##0.00'; PCT='0.0%'

wb=Workbook(); ws=wb.active; ws.title="Comparativo"
ws.sheet_view.showGridLines=False
for col,w in {"A":5,"B":44,"C":10,"D":14,"E":14,"F":11,"G":16,"H":34}.items():
    ws.column_dimensions[col].width=w

def merge_set(rng,value,font=None,fill=None,align=None,fmt=None,border=False):
    ws.merge_cells(rng); c=ws[rng.split(":")[0]]; c.value=value
    if font:c.font=font
    if align:c.alignment=align
    if fmt:c.number_format=fmt
    mc,mr,xc,xr=range_boundaries(rng)
    for r in range(mr,xr+1):
        for cc in range(mc,xc+1):
            cel=ws.cell(row=r,column=cc)
            if fill:cel.fill=PatternFill("solid",fgColor=fill)
            if border:cel.border=BORDER
    return c

# ---- Encabezado ------------------------------------------------------------
merge_set("A1:H1","COMPARATIVO DE PRECIOS — MURO PERIMETRAL",F(16,True,AZUL),align=CENTER)
ws.row_dimensions[1].height=26
merge_set("A2:H2","Presupuesto del 30-jul-2026  vs.  versión actual (04-ago-2026) · Mercado de ferreterías, Managua · Precios con IVA 15% incluido",
          F(11,False,"404040",italic=True),align=CENTER)
ws.row_dimensions[2].height=20

# ---- Tabla 1: precios unitarios -------------------------------------------
merge_set("A4:H4","1.  COMPARATIVO DE PRECIOS UNITARIOS (C$, IVA INCLUIDO)",F(12,True,AZUL),
          fill="DDEBF7",align=Alignment("left",vertical="center"),border=True)
ws.row_dimensions[4].height=24
heads=["N°","Descripción","Unidad","P. Unit.\n30-jul","P. Unit.\nActual","Variación","Tendencia","Observación"]
for i,h in enumerate(heads):
    c=ws.cell(row=5,column=1+i,value=h); c.font=F(11,True,"FFFFFF")
    c.fill=PatternFill("solid",fgColor=AZUL); c.alignment=CENTER; c.border=BORDER
ws.row_dimensions[5].height=38

# (descripción, unidad, precio_jul, precio_actual, observación)
items=[
 ("Bloque de concreto de 6\" (15x20x40 cm)","Unidad",40,22,"Ref. CONCRENIC/ARCON"),
 ("Varilla corrugada de 3/8\" G-40","Unidad",1700,185,"30-jul: 19 u @ C$1,700 (precio atípico); Actual: 79 u @ C$185. Validar."),
 ("Estribo de 1/4\" de 10x10 cm","Unidad",22,15,""),
 ("Alambre de amarre recocido","Libra",48,28,"Ref. Padua ≈C$2,050/qq. 30-jul cal.16 / Actual cal.18"),
 ("Cemento gris tipo GU (bolsa 42.5 kg)","Bolsa",465,395,"Ref. SINSA ≈C$385–400"),
 ("Arena de construcción","m³",950,700,"Ref. mercado ≈C$600–800"),
 ("Piedrín de 1/2\"","m³",1300,1000,"Ref. mercado C$900–1,000"),
 ("Tabla de pino de 1x12x6 varas","Unidad",320,500,"Actual más alto"),
 ("Regla de pino de 1x2x6 varas","Unidad",110,110,""),
 ("Clavos corrientes de 2 1/2\"","Libra",48,35,""),
 ("Clavos de acero de 2 1/2\"","Docena",70,55,""),
 ("Adhesivo Plaster Bond","Litro",185,230,"Actual más alto"),
 ("Rollo de concertina / serpentina","Rollo",950,2200,"Mercado usual C$1,800–2,500; validar el de 30-jul"),
 ("Grapas","Libra",55,35,""),
 ("Electrodo soldadura 6011 3/32\" (caja)","Caja",520,780,"Actual más alto"),
 ("Pintura anticorrosiva","Cuarto",190,180,"30-jul: C$380 el juego de 2 cuartos (≈C$190 c/u)"),
 ("Diluyente (thinner)","Galón",320,400,"Actual más alto"),
 ("Tubo cuadrado 1 1/2\" chapa 14 (6 m)","Unidad",480,700,"Ref. SOFECONSA: 1\"x1\" ch14 ≈C$420+IVA"),
 ("Lámina negra de 4x8 pies","Unidad",1050,1250,"30-jul: cal.14 (~2 mm) / Actual: 1 mm (cal.20)"),
 ("Par de bisagra de bala","Par",220,120,""),
 ("Pasador (aldaba) de 5\"","Unidad",95,90,""),
 ("Disco de corte fino de 7\"","Unidad",280,75,""),
]

var_cache={}
row=6
for i,(desc,uni,pj,pa,obs) in enumerate(items,start=1):
    ws.cell(row=row,column=1,value=i).font=F(10)
    ws.cell(row=row,column=2,value=desc).font=F(10)
    ws.cell(row=row,column=3,value=uni).font=F(10)
    cj=ws.cell(row=row,column=4,value=pj); cj.font=F(10); cj.number_format=MON
    ca=ws.cell(row=row,column=5,value=pa); ca.font=F(10); ca.number_format=MON
    cv=ws.cell(row=row,column=6,value=f"=(E{row}-D{row})/D{row}"); cv.font=F(10); cv.number_format=PCT
    var_cache[f"F{row}"]=round((pa-pj)/pj,4)
    # tendencia (etiqueta derivada, snapshot)
    if pa<pj:  txt,col="▼ más barato",VERDE
    elif pa>pj:txt,col="▲ más caro",ROJO
    else:      txt,col="= igual",GRIST
    ct=ws.cell(row=row,column=7,value=txt); ct.font=F(10,True,col)
    co=ws.cell(row=row,column=8,value=obs); co.font=F(9,italic=True,color="595959")
    ws.cell(row=row,column=1).alignment=CENTER
    ws.cell(row=row,column=2).alignment=LEFT
    ws.cell(row=row,column=3).alignment=CENTER
    ws.cell(row=row,column=4).alignment=RIGHT
    ws.cell(row=row,column=5).alignment=RIGHT
    ws.cell(row=row,column=6).alignment=RIGHT
    ws.cell(row=row,column=7).alignment=CENTER
    ws.cell(row=row,column=8).alignment=LEFTT
    for cc in range(1,9): ws.cell(row=row,column=cc).border=BORDER
    ws.row_dimensions[row].height=30 if len(obs)>34 else 20
    row+=1
fin1=row  # primera libre

# ---- Tabla 2: totales ------------------------------------------------------
t2=fin1+1
merge_set(f"A{t2}:H{t2}","2.  COMPARATIVO DE TOTALES DEL PRESUPUESTO (C$)",F(12,True,AZUL),
          fill="DDEBF7",align=Alignment("left",vertical="center"),border=True)
ws.row_dimensions[t2].height=24
h2=t2+1
for col,txt in [("A","Concepto"),("E","30-jul-2026"),("F","Versión actual"),("G","Diferencia"),("H","Diferencia %")]:
    pass
# encabezado tabla 2 (Concepto combinado A:D)
merge_set(f"A{h2}:D{h2}","Concepto",F(11,True,"FFFFFF"),fill=AZUL,align=CENTER,border=True)
for col,txt in [("E","30-jul-2026"),("F","Versión actual"),("G","Diferencia"),("H","Diferencia %")]:
    c=ws[f"{col}{h2}"]; c.value=txt; c.font=F(11,True,"FFFFFF")
    c.fill=PatternFill("solid",fgColor=AZUL); c.alignment=CENTER; c.border=BORDER
ws.row_dimensions[h2].height=32

# filas: (concepto, valor_jul, valor_act) ; diferencia y % con fórmula
tot_rows=[
 ("Subtotal materiales (IVA incluido)",51217,35387,False),
 ("Imprevistos 5%",3335.85,1769.35,False),
 ("Subtotal mano de obra",15500,15500,False),
 ("TOTAL GENERAL",70052.85,52656.35,True),
]
tot_cache={}
rr=h2+1
for concepto,vj,va,es_total in tot_rows:
    fill=GRIS_TOT if es_total else GRIS_SUB
    blanco=es_total
    merge_set(f"A{rr}:D{rr}",concepto,F(12 if es_total else 11,True,"FFFFFF" if blanco else "000000"),
              fill=fill,align=RIGHT,border=True)
    ce=ws[f"E{rr}"]; ce.value=vj
    cf=ws[f"F{rr}"]; cf.value=va
    cg=ws[f"G{rr}"]; cg.value=f"=F{rr}-E{rr}"
    ch=ws[f"H{rr}"]; ch.value=f"=IFERROR((F{rr}-E{rr})/E{rr},0)"
    tot_cache[f"G{rr}"]=round(va-vj,2)
    tot_cache[f"H{rr}"]=round((va-vj)/vj,4) if vj else 0
    for cc,cell in ((5,ce),(6,cf),(7,cg)):
        cell.number_format=MON
    ch.number_format=PCT
    for col in ("E","F","G","H"):
        cel=ws[f"{col}{rr}"]
        cel.font=F(12 if es_total else 11,True,"FFFFFF" if blanco else "000000")
        cel.alignment=RIGHT; cel.border=BORDER
        cel.fill=PatternFill("solid",fgColor=fill)
    ws.row_dimensions[rr].height=26 if es_total else 22
    rr+=1
# nota base imprevistos
merge_set(f"A{rr}:H{rr}",
          "Nota: el 30-jul los imprevistos (5%) se calcularon sobre materiales + mano de obra (base C$66,717); "
          "en la versión actual se calculan solo sobre materiales (base C$36,387), según tu indicación.",
          F(9,italic=True,color="595959"),align=LEFTT)
ws.row_dimensions[rr].height=30
rr+=1

# ---- Veredicto -------------------------------------------------------------
vr=rr+1
merge_set(f"A{vr}:H{vr}","VEREDICTO",F(12,True,AZUL),fill="F2F2F2",
          align=Alignment("left",vertical="center"),border=True)
ws.row_dimensions[vr].height=22
veredicto=[
 "• En conjunto NO: la versión actual es MÁS BARATA. El total bajó de C$ 70,052.85 (30-jul) a C$ 52,656.35, una reducción de C$ 17,396.50 (−24.8%). Los precios actuales están contrastados con el mercado nicaragüense (SINSA, SOFECONSA, Padua, tecnoguias).",
 "• En precios unitarios, la versión actual es más barata en 15 de 22 materiales; 7 subieron (tabla de pino, Plaster Bond, concertina, soldadura, thinner, tubo y lámina) y 1 quedó igual (regla).",
 "• El mayor efecto viene de la varilla de 3/8\": el 30-jul se registró a C$ 1,700 por unidad —precio atípico, pues una varilla ronda C$ 180–200— frente a C$ 185 en la versión actual. Además cambió la cantidad (19 vs 79 u); conviene confirmar cuál es la correcta según la lista.",
 "• Puntos a validar con cotización formal: varilla de 3/8\" (precio y cantidad), rollo de concertina y el calibre de la lámina del portón (cal.14 vs 1 mm).",
]
rv=vr+1
for t in veredicto:
    merge_set(f"A{rv}:H{rv}",t,F(10),align=LEFTT,border=True)
    lineas=max(1,int(len(t)/135)+1)
    ws.row_dimensions[rv].height=15*lineas+8
    rv+=1

# ---- Leyenda ---------------------------------------------------------------
lg=rv+1
merge_set(f"A{lg}:H{lg}",
          "Leyenda:  ▼ más barato (verde) = el precio actual es menor que el del 30-jul   |   ▲ más caro (rojo) = el precio actual es mayor.  "
          "Precios de referencia de mercado (IVA incluido); no constituyen cotización formal de proveedor.",
          F(9,italic=True,color="595959"),align=LEFTT)
ws.row_dimensions[lg].height=30

# ---- Impresión -------------------------------------------------------------
ws.freeze_panes="A6"
ws.print_area=f"A1:H{lg}"
ws.print_title_rows="5:5"
ws.page_setup.orientation="landscape"
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
ws.sheet_properties.pageSetUpPr.fitToPage=True
ws.page_margins.left=ws.page_margins.right=0.4
ws.page_margins.top=ws.page_margins.bottom=0.5
ws.print_options.horizontalCentered=True

try: wb.calculation.fullCalcOnLoad=True
except Exception as e: print("aviso:",e)

ARCH="Comparativo_Precios_Muro.xlsx"
wb.save(ARCH)

# ---- Inyección de valores en caché ----------------------------------------
import zipfile,re,os
cache={}; cache.update(var_cache); cache.update(tot_cache)
with zipfile.ZipFile(ARCH) as z:
    names=z.namelist(); blobs={n:z.read(n) for n in names}
xml=blobs["xl/worksheets/sheet1.xml"].decode("utf-8")
for coord,val in cache.items():
    sval=str(int(val)) if float(val)==int(val) else f"{val:.4f}"
    pat=re.compile(rf'(<c r="{coord}"[^>]*><f>.*?</f>)<v\s*/>(</c>)',re.DOTALL)
    xml,n=pat.subn(rf'\g<1><v>{sval}</v>\g<2>',xml)
    if n==0: print("WARN sin match:",coord)
blobs["xl/worksheets/sheet1.xml"]=xml.encode("utf-8")
tmp=ARCH+".tmp"
with zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED) as zo:
    for n in names: zo.writestr(n,blobs[n])
os.replace(tmp,ARCH)
print("OK comparativo. filas tabla1:",6,"->",fin1-1," tabla2 header:",h2)
